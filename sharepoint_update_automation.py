from office365.sharepoint.client_context import ClientContext
from configparser import ConfigParser
import datetime
from os.path import getsize
from openpyxl import load_workbook
import re

# load auth credentials & settings from settings.cfg file
def load_setting(specified_setting):
    cp = ConfigParser()
    cp.read('./automation/settings.cfg')
    return dict(cp.items(specified_setting))   

def download_daily_data(): 
    curr_date = datetime.date.today()

    # load sharepoint credentials and info
    sharepoint_site, folder_url = load_setting('sharepoint_testing_path').values()
    ctx = ClientContext(sharepoint_site).with_client_certificate(**load_setting('sharepoint_cert_credentials'))

    # construct path to file by adding to datafolder_url
    file_url = folder_url + curr_date.strftime('%Y/%m_%B/%y%m%d Daily Data.xlsx') # format for file path: yyyy/mm_MonthStr/yymmdd Daily Data.xlsx             
    # construct path for local filename to download as
    download_path = curr_date.strftime('./data_sources/daily_data/daily data %Y%m%d.xlsx')

    ## TODO FIX THIS TO ACTUALLY WORK AND DETERMINE IF FILE EXISTS BEFORE ATTEMPTING TO DOWNLOAD
    folder = ctx.web.get_file_by_server_relative_url(file_url).select("Exists").get().execute_query()
    if folder.exists:
        print("File is found")
    else:
        print("File not found")

    print(file_url)
    with open(download_path, "wb") as local_file:
        print('Downloading today\'s daily data')
        for i in range(5):
            try:
                [print(f'Attempt {i+1}/5') if i > 0 else ''][0]
                ctx.web.get_file_by_server_relative_url(file_url).download_session(local_file, lambda x: print(f'Downloaded {x/1e6:.2f} MB'),chunk_size=int(5e6)).execute_query()
                break
            except:
                print('Download error...trying again')
                if i == 4:
                    print('Daily data download error')
                    return False
    if getsize(download_path) > 40000:  # check that downloaded filesize is > 40 KB, in case of any download error
        print(f'Daily data successfully downloaded to {download_path}')
        return download_path
    else:
        print('Daily data download error')
        return False
    
file_path = download_daily_data()

wb = load_workbook(filename = file_path, data_only=True) # use data_only to get calculated values, otherwise formulas will be returned

ws = wb['Scorecard']

daily_data = {}
for row in ws.iter_rows(min_row=2, max_row=80, values_only=True):
    if row[1] != None:
        temp_data_list = list(row[1:]) # convert to a list for row data to be mutable (since iter_rows returns a tuple)
        for idx, n in enumerate(temp_data_list):
            # catch & correct data entry error with a decimal point repeated more than once (will be a string in this case, so check type to detect, though some strings are valid, but this cuts down on unneccesary processing)
            if type(n) == str:
                print('before', n)
                # use regular expression to replace zero or more occurances of a number, followed by two or more decimals, followed by zero or more occurances of a number, with the extra decimals removed
                n = re.sub(r'([0-9]*)(\.{2,})([0-9]*)',r'\1.\3',n)
                print('after', n)
                # convert to numeric & update data list using try/except (assume failure to convert to float means this value is meant to be a string i.e., a text comment)
                try:
                    temp_data_list[idx] = float(n)
                    print('value updated')
                except:
                    print('error, cannot convert, not updating value')
                    pass
                print('after2', n)
        print(row[0], temp_data_list)
        daily_data[str(row[0])] = temp_data_list
        
print(daily_data)

scorecard_wb = load_workbook(filename = './data_sources/test_scorecard_data.xlsx')

curr_date = datetime.date.today()
print(curr_date)

###############################
## Update scorecard workbook ## 
##    with daily data        ## 
###############################

# first iterate through the daily data values
for idx, (pond_name, daily_data_pond_values) in enumerate(daily_data.items()):
    print(f'-------------------\nUpdating scorecard data for {pond_name}')
    # use try/except clause to catch error when the pond_name is not present in the scorecard workbook
    try:
        pond_ws = scorecard_wb[pond_name]
    except:
        print(f'ERROR: could not find sheet named "{pond_name}"')
        # check for edge case where "pond_name (media only)" refers to "pond_name HRP" on the scorecard workbook
        pond_name_split = pond_name.split()
        if len(pond_name_split) == 3 and (pond_name_split[1], pond_name_split[2]) == ('(media', 'only)'): 
            n_pond_name = f'{pond_name_split[0]} HRP'
            print(f'Trying sheet "{n_pond_name}" instead') 
            try:
                pond_ws = scorecard_wb[n_pond_name]
            except:
                print(f'ERROR: could not find sheet named "{n_pond_name}"')
                continue
        else:       
            continue
    
    #iterate through the corresponding scorecard workbook pond sheet to find the row for current date
    row_idx = None # set row_idx (index for the current date row in scorecard worksheet) to None to catch case when date doesn't exist in worksheet
    for idx2, row in enumerate(pond_ws.iter_rows(min_row=2, values_only=True), start=2):
        # keep track of the last nonempty row (for case when the date isn't present in the scorecard worksheet and needs to be added)
        if row[0] != None:
            last_nonempty_row_idx = idx2
            last_nonempty_row_date = None # set to none to reset val on each row iteration, in case this row doesn't contain a valid date
            if isinstance(row[0], datetime.datetime):
                last_nonempty_row_date = row[0].date()
        
        # check if first column in row is a datetime value, and if it is the current date
        # if this row is for the current date, set row_idx and break the loop
        if isinstance(row[0], datetime.datetime) and row[0].date() == curr_date:
            row_idx = idx2
            #print(f'Found date on row {row_idx}')
            break
    
    # case where current date isn't present in the scorecard worksheet, row_idx will be None if prev loop iterated through all rows without finding it
    if row_idx == None: 
        # case when the last nonempty row was a valid date
        if last_nonempty_row_date != None:
            diff_days = (curr_date - last_nonempty_row_date).days
            if diff_days <= 7: # fill-in missing row dates if the difference between current and last is 7 or fewer days 
                row_idx = last_nonempty_row_idx + diff_days # set the row index for current date
                fill_row_idx = last_nonempty_row_idx+1
                fill_row_date = last_nonempty_row_date + datetime.timedelta(days=1)
                for _ in range(diff_days-1):
                    pond_ws[f'A{fill_row_idx}'].value = fill_row_date
                    pond_ws[f'A{fill_row_idx}'].number_format = 'm/d/yyyy'
                    fill_row_idx += 1
                    fill_row_date += datetime.timedelta(days=1)
            else: # case when more than 7 days between current and last nonempty row, use blank rows as space between
                row_idx = last_nonempty_row_idx + 3 # set row idx to the last nonempty row + 3 (keep 2 empty rows as a space between)
        else: # case when the last nonempty row is not a valid date
            row_idx = last_nonempty_row_idx + 3 # set row idx to the last nonempty row + 3 (keep 2 empty rows as a space between)
        
        # set the current date row value
        pond_ws[f'A{row_idx}'].value = curr_date
        pond_ws[f'A{row_idx}'].number_format = 'm/d/yyyy'
        
    # update scorecard worksheet with new data
    update_range = pond_ws[f'B{row_idx}:U{row_idx}'][0] # a tuple containing the cells to be updated, column B through U for specified row index
    for col_idx, cell in enumerate(update_range):
        cell.value = daily_data_pond_values[col_idx] # daily_data_pond_values is corresponding to the daily_data dict for current sheet

    print(f'Successfully updated scorecard data for {pond_name}')
    
scorecard_wb.save('test_.xlsx')        
print('Updated excel file successfully saved')